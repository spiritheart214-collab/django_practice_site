import datetime
import csv

import openpyxl
from openpyxl.styles import Font, Alignment
from django.contrib import admin
from django.db.models import QuerySet
from django.db.models.options import Options
from django.http import HttpRequest, HttpResponse


from django.contrib import admin
class ExportAsCSVMixin:
    @admin.action(description="Export as CSV")
    def export_csv(self: admin.ModelAdmin, request: HttpRequest, queryset: QuerySet):
        meta: Options = self.model._meta
        fields_name = [field.name for field in meta.fields]

        # Преобразуем model_name в читаемый формат
        model_name_readable = meta.verbose_name_plural or meta.model_name
        filename = f"{model_name_readable}-export.csv"  # ✅

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f"attachment; filename={filename}"

        csv_writer = csv.writer(response)
        csv_writer.writerow(fields_name)

        for obj in queryset:
            csv_writer.writerow([getattr(obj, field) for field in fields_name])

        return response


class ExportAsExcelMixin:
    @admin.action(description="Export as Excel")
    def export_excel(self: admin.ModelAdmin, request: HttpRequest, queryset: QuerySet):
        """Экспортирует выбранные объекты в Excel файл."""

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"

        meta = self.model._meta
        field_names = [field.name for field in meta.fields]

        # Заголовки
        for col_num, field_name in enumerate(field_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = field_name
            cell.font = Font(bold=True)

        # Данные
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(field_names, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = getattr(obj, field_name)

                # Простая обработка значений
                if value is None:
                    value = ""
                elif hasattr(value, '__call__'):
                    value = value()
                elif hasattr(value, 'strftime'):  # datetime объекты
                    # Преобразуем в строку чтобы избежать проблем с timezone
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    value = "Yes" if value else "No"

                cell.value = value

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={meta.model_name}-export.xlsx"

        workbook.save(response)
        return response
